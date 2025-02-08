
# PyQt5 imports
from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QApplication, QVBoxLayout, QHBoxLayout, 
    QLabel, QPushButton, QFileDialog, QMessageBox, QGroupBox,
    QFormLayout, QLineEdit, QTextEdit, QComboBox, QDateEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar,
    QCheckBox, QDialog, QStatusBar, QToolBar, QAction, QTabWidget, QGridLayout, QDialogButtonBox
)
from PyQt5.QtCore import Qt, QDate, QTimer, QSize
from PyQt5.QtGui import QIcon, QPixmap, QFont, QColor
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis

# Deep learning
import tensorflow as tf 
from tensorflow.keras.models import load_model
from ultralytics import YOLO
import tempfile
from io import BytesIO
# Image processing
import cv2
import sqlite3
import numpy as np

# System
import os
import sys
from datetime import datetime, timedelta
import shutil
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
# Database
from database import create_connection, backup_database, restore_database
import time
class ThemeManager:
    def __init__(self):
        self.dark_theme = {
            'bg': '#2c3e50',
            'text': '#ecf0f1',
            'accent': '#3498db', 
            'success': '#2ecc71',
            'warning': '#f1c40f',
            'error': '#e74c3c'
        }
        
        self.light_theme = {
            'bg': '#f5f6fa',
            'text': '#2c3e50',
            'accent': '#3498db',
            'success': '#27ae60', 
            'warning': '#f39c12',
            'error': '#c0392b'
        }
        
        self.current_theme = 'light'
        
    def get_theme(self):
        return self.light_theme if self.current_theme == 'light' else self.dark_theme
        
    def toggle_theme(self):
        self.current_theme = 'dark' if self.current_theme == 'light' else 'light'
class MainWindow(QMainWindow):
   def __init__(self, role):
        super().__init__()
        self.role = role
        self.theme = ThemeManager()
        
        # Initialize key attributes
        self.current_page = 1
        self.total_patients = 0
        
        # Initialize UI components
        self.init_window()
        self.init_folders()
        self.setup_models()
        self.init_ui()
        
        # Load data
        self.load_data()
        
        # Auto update stats timer
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_statistics)
        self.update_timer.start(300000)  # Update every 5 minutes
        # self.reports_dir = os.path.abspath('./Code_app/reports')

   def init_window(self):
       """Initialize main window properties"""
       self.setWindowTitle('Breast Cancer Diagnosis System')
       self.setGeometry(100, 100, 1280, 800) 
       self.setMinimumSize(1000, 600)
       self.setWindowIcon(QIcon('./Code_app/assets/logo.png'))

   def init_folders(self):
       """Create required folders if not exist"""
       folders = {
           'processed': './Code_app/processed',
           'diagnoses': './Code_app/diagnoses', 
           'reports': './Code_app/reports',
           'backup': './Code_app/backup'
       }
       
       for name, path in folders.items():
           if not os.path.exists(path):
               os.makedirs(path)
           setattr(self, f'{name}_dir', path)
   def toggle_theme(self):
        """Toggle between light and dark themes"""
        self.theme.toggle_theme()
        self.apply_theme()
   def setup_models(self):
       """Load and configure AI models"""
       try:
           # Load YOLO detection model
           self.yolo_model = YOLO('./Code_app/models/best.pt')
           
           # Load and compile classification model
           self.keras_model = load_model('./Code_app/models/DenseNet121_model.h5')
           self.keras_model.compile(
               optimizer='adam',
               loss='categorical_crossentropy',
               metrics=['accuracy']
           )
           
           # Load model metrics or use defaults
           metrics_file = './Code_app/models/metrics.json'
           if os.path.exists(metrics_file):
               with open(metrics_file) as f:
                   self.model_metrics = json.load(f)
           else:
               self.model_metrics = {
                   'detect': {'precision': 0, 'recall': 0, 'map50': 0},
                   'classify': {'accuracy': 0, 'f1': 0, 'auc': 0}
               }
               
       except Exception as e:
           QMessageBox.critical(self, 'Error', f'Cannot load models: {str(e)}')
           sys.exit(1)

   def clear_form(self):
        """Clear all patient form fields"""
        self.patient_name.clear()
        self.patient_dob.setDate(QDate.currentDate())
        self.patient_gender.setCurrentIndex(0)
        self.patient_id_number.clear()
        self.patient_address.clear()
        self.patient_phone.clear()
        self.patient_email.clear()
        self.patient_table.clearSelection()
        self.validate_form()

   def init_ui(self):
       """Initialize user interface"""
       # Central widget
       self.central_widget = QWidget()
       self.setCentralWidget(self.central_widget)
       self.main_layout = QVBoxLayout(self.central_widget)
       self.main_layout.setContentsMargins(20, 20, 20, 20)
       
       # Create toolbar
       self.create_toolbar()
       
       # Create tabs
       self.tabs = QTabWidget()
       self.tabs.setDocumentMode(True)
       
       # Add tabs
       self.patient_tab = self.create_patient_tab()
       self.tabs.addTab(self.patient_tab, "Patient Management")
       
       self.diagnosis_tab = self.create_diagnosis_tab()
       self.tabs.addTab(self.diagnosis_tab, "Diagnosis")
       
       self.reports_tab = self.create_reports_tab()
       self.tabs.addTab(self.reports_tab, "Reports")
       
       if self.role == 'admin':
           self.admin_tab = self.create_admin_tab()
           self.tabs.addTab(self.admin_tab, "Admin")
           
       self.main_layout.addWidget(self.tabs)
       
       # Status bar
       self.status_bar = QStatusBar()
       self.setStatusBar(self.status_bar)
       
       # Apply theme
       self.apply_theme()

   def create_toolbar(self):
       """Create main toolbar"""
       toolbar = QToolBar()
       toolbar.setMovable(False)
       toolbar.setIconSize(QSize(24, 24))
       self.addToolBar(toolbar)
       
       # Theme toggle button
       theme_btn = QAction(QIcon('./Code_app/assets/theme.png'), 'Toggle Theme', self)
       theme_btn.triggered.connect(self.toggle_theme)
       theme_btn.setStatusTip('Toggle light/dark theme')
       toolbar.addAction(theme_btn)
       
       toolbar.addSeparator()
       
       # Backup/Restore buttons
       backup_btn = QAction(QIcon('./Code_app/assets/backup.png'), 'Backup Data', self)
       backup_btn.triggered.connect(lambda: backup_database(self))
       backup_btn.setStatusTip('Backup database')
       toolbar.addAction(backup_btn)
       
       restore_btn = QAction(QIcon('./Code_app/assets/restore.png'), 'Restore Data', self)
       restore_btn.triggered.connect(lambda: restore_database(self))
       restore_btn.setStatusTip('Restore from backup')
       toolbar.addAction(restore_btn)
       
       toolbar.addSeparator()
       
       # Model info button
       model_btn = QAction(QIcon('./Code_app/assets/model.png'), 'Model Information', self)
       model_btn.triggered.connect(self.show_model_info)
       model_btn.setStatusTip('View model information')
       toolbar.addAction(model_btn)
    
       export_excel_btn = QAction(QIcon('./Code_app/assets/excel.png'), 'Export Excel', self)
       export_excel_btn.triggered.connect(self.export_excel)
       export_excel_btn.setStatusTip('Export patient list to Excel')
       toolbar.addAction(export_excel_btn)

   def apply_theme(self):
        """Apply theme to entire application"""
        theme = self.theme.get_theme()
        
        # Global stylesheet
        self.setStyleSheet(f"""
            QWidget {{
                background-color: {theme['bg']};
                color: {theme['text']};
                font-family: Arial;
            }}
            
            QTabWidget::pane {{
                border: 1px solid {theme['accent']};
                background: {theme['bg']};
                border-radius: 5px;
            }}
            
            QTabBar::tab {{
                background: {theme['bg']};
                color: {theme['text']};
                padding: 12px 20px;
                border: 1px solid {theme['accent']};
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                min-width: 120px;
                font-weight: bold;
            }}
            
            QTabBar::tab:selected {{
                background: {theme['accent']};
                color: white;
            }}
            
            QGroupBox {{
                background-color: {theme['bg']};
                color: {theme['text']};
                border: 1px solid {theme['accent']};
                border-radius: 8px;
                padding: 15px;
                margin-top: 1.5em;
                font-weight: bold;
            }}
            
            QPushButton {{
                background-color: {theme['accent']};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                min-width: 100px;
                font-weight: bold;
            }}
            
            QPushButton:hover {{
                background-color: {theme['accent'] + '99'};
                border: 1px solid {theme['accent']};
            }}
            
            QPushButton:disabled {{
                background-color: {theme['bg']};
                color: {theme['text'] + '77'};
                border: 1px solid {theme['text'] + '44'};
            }}
            
            QLineEdit, QTextEdit, QComboBox {{
                background: {theme['bg']};
                color: {theme['text']};
                border: 1px solid {theme['accent']};
                border-radius: 4px;
                padding: 8px;
                selection-background-color: {theme['accent']};
            }}
            
            QComboBox::drop-down {{
                border: none;
                width: 20px;
            }}
            
            QComboBox::down-arrow {{
                image: url(./Code_app/assets/arrow-down.png);
                width: 12px;
                height: 12px;
            }}
            
            QTableWidget {{
                background: {theme['bg']};
                color: {theme['text']};
                gridline-color: {theme['accent']};
                border: none;
                selection-background-color: {theme['accent'] + '44'};
            }}
            
            QHeaderView::section {{
                background-color: {theme['accent']};
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }}
            
            QScrollBar:vertical {{
                background: {theme['bg']};
                width: 12px;
                border-radius: 6px;
            }}
            
            QScrollBar::handle:vertical {{
                background: {theme['accent'] + '66'};
                min-height: 30px;
                border-radius: 6px;
            }}
            
            QScrollBar::up-arrow:vertical,
            QScrollBar::down-arrow:vertical,
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {{
                border: none;
                background: none;
            }}
            
            QProgressBar {{
                border: 1px solid {theme['accent']};
                border-radius: 4px;
                background: {theme['bg']};
                text-align: center;
                padding: 2px;
            }}
            
            QProgressBar::chunk {{
                background: {theme['accent']};
                border-radius: 3px;
            }}
        """)
   def create_patient_tab(self):
        """Create patient management tab"""
        tab = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Patient Info Form
        form_group = QGroupBox("Patient Information")
        form_layout = QFormLayout()

        # Input fields
        self.patient_name = QLineEdit()
        self.patient_name.setPlaceholderText("Enter patient name...")
        self.patient_name.textChanged.connect(self.validate_form)

        self.patient_dob = QDateEdit()
        self.patient_dob.setFixedHeight(30)
        self.patient_dob.setCalendarPopup(True)
        self.patient_dob.setDisplayFormat("dd/MM/yyyy")
        self.patient_dob.setMinimumDate(QDate(1900, 1, 1))
        self.patient_dob.setMaximumDate(QDate.currentDate())

        self.patient_gender = QComboBox()
        self.patient_gender.addItems(['Male', 'Female'])

        self.patient_id_number = QLineEdit()
        self.patient_id_number.setPlaceholderText("Enter patient ID...")
        self.patient_id_number.setInputMask("999999999999;_")
        self.patient_id_number.textChanged.connect(self.validate_form)

        self.patient_address = QTextEdit()
        self.patient_address.setPlaceholderText("Enter patient address...")
        self.patient_address.setMaximumHeight(60)

        self.patient_phone = QLineEdit()
        self.patient_phone.setPlaceholderText("Enter patient phone...")
        self.patient_phone.setInputMask("9999999999;_")
        self.patient_phone.textChanged.connect(self.validate_form)

        self.patient_email = QLineEdit()
        self.patient_email.setPlaceholderText("Enter patient email...")

        # Add fields to form
        fields = [
            ("Name:", self.patient_name),
            ("Date of Birth:", self.patient_dob),
            ("Gender:", self.patient_gender),
            ("ID Number:", self.patient_id_number),
            ("Address:", self.patient_address),
            ("Phone Number:", self.patient_phone),
            ("Email:", self.patient_email)
        ]

        for label, widget in fields:
            form_layout.addRow(QLabel(label), widget)

        form_group.setLayout(form_layout)

        # Action Buttons
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("Add Patient")
        self.add_btn.setIcon(QIcon("./Code_app/assets/add.png"))
        self.add_btn.clicked.connect(self.add_patient)
        self.add_btn.setEnabled(False)

        self.update_btn = QPushButton("Update")
        self.update_btn.setIcon(QIcon("./Code_app/assets/update.png")) 
        self.update_btn.clicked.connect(self.update_patient)
        self.update_btn.setEnabled(False)

        self.clear_btn = QPushButton("Clear Form")
        self.clear_btn.setIcon(QIcon("./Code_app/assets/clear.png"))
        self.clear_btn.clicked.connect(self.clear_form)

        self.delete_btn = QPushButton("Delete")
        self.delete_btn.setIcon(QIcon("./Code_app/assets/delete.png"))
        self.delete_btn.clicked.connect(self.delete_patient)
        self.delete_btn.setEnabled(False)

        btn_layout.addStretch()
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.update_btn)
        btn_layout.addWidget(self.clear_btn)
        btn_layout.addWidget(self.delete_btn)
        # Search section
        search_group = QGroupBox("Advanced Search")
        search_layout = QGridLayout()

        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name, ID, or phone number...")
        self.search_input.textChanged.connect(self.search_patient)

        # Filters
        self.age_range = QComboBox()
        self.age_range.addItems(['All', '0-18', '19-40', '41-60', 'Over 60'])
        self.age_range.currentTextChanged.connect(self.search_patient)

        self.gender_filter = QComboBox()
        self.gender_filter.addItems(['All', 'Male', 'Female'])
        self.gender_filter.currentTextChanged.connect(self.search_patient)

        # Add to layout  
        search_layout.addWidget(QLabel("Keyword:"), 0, 0)
        search_layout.addWidget(self.search_input, 0, 1, 1, 3)
        search_layout.addWidget(QLabel("Age Range:"), 1, 0)
        search_layout.addWidget(self.age_range, 1, 1)
        search_layout.addWidget(QLabel("Gender:"), 1, 2)
        search_layout.addWidget(self.gender_filter, 1, 3)

        search_group.setLayout(search_layout)

        # Patient Table
        table_group = QGroupBox("Patient List")
        table_layout = QVBoxLayout()

        self.patient_table = QTableWidget()
        self.patient_table.setColumnCount(8)
        self.patient_table.setHorizontalHeaderLabels([
            'ID', 'Name', 'Date of Birth', 'Gender',   
            'ID Card/CCCD', 'Address', 'Phone Number', 'Email'
        ])

        # Table settings
        header = self.patient_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionResizeMode(0, QHeaderView.Fixed)  
        header.setSectionResizeMode(2, QHeaderView.Fixed)  
        header.setSectionResizeMode(3, QHeaderView.Fixed)  
        header.setSectionResizeMode(4, QHeaderView.Fixed)  
        header.setSectionResizeMode(6, QHeaderView.Fixed) 

        self.patient_table.setColumnWidth(0, 50)
        self.patient_table.setColumnWidth(2, 100)
        self.patient_table.setColumnWidth(3, 80)
        self.patient_table.setColumnWidth(4, 120)
        self.patient_table.setColumnWidth(6, 100)

        self.patient_table.verticalHeader().setVisible(False)
        self.patient_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.patient_table.setSelectionMode(QTableWidget.SingleSelection)
        self.patient_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.patient_table.clicked.connect(self.patient_table_clicked)
        # Pagination
        pagination_layout = QHBoxLayout()
        
        self.page_size = QComboBox()
        self.page_size.addItems(['10', '25', '50', '100'])
        self.page_size.currentTextChanged.connect(self.change_page_size)
        
        self.prev_btn = QPushButton("Prev")
        self.prev_btn.setIcon(QIcon("./Code_app/assets/prev.png"))
        self.prev_btn.clicked.connect(lambda: self.change_page(-1))
        self.prev_btn.setEnabled(False)
        
        self.page_label = QLabel("Page 1/1")
        
        self.next_btn = QPushButton("Next")
        self.next_btn.setIcon(QIcon("./Code_app/assets/next.png"))
        self.next_btn.clicked.connect(lambda: self.change_page(1))
        self.next_btn.setEnabled(False)
        
        pagination_layout.addWidget(QLabel("Number of rows:"))
        pagination_layout.addWidget(self.page_size)
        pagination_layout.addStretch()
        pagination_layout.addWidget(self.prev_btn)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(self.next_btn)
        
        table_layout.addWidget(self.patient_table)
        table_layout.addLayout(pagination_layout)
        table_group.setLayout(table_layout)
        
        # Add all components to main layout
        layout.addWidget(form_group)
        layout.addLayout(btn_layout)  
        layout.addWidget(search_group)
        layout.addWidget(table_group)
        
        tab.setLayout(layout)
        return tab

   def validate_form(self):
        """Form validation"""
        name = self.patient_name.text().strip()
        phone = self.patient_phone.text().strip()
        id_number = self.patient_id_number.text().strip()
        
        has_required = bool(name and phone and len(phone) == 10 and id_number)
        selected_row = self.patient_table.currentRow()
        
        self.add_btn.setEnabled(has_required and selected_row < 0)
        self.update_btn.setEnabled(has_required and selected_row >= 0)
        self.delete_btn.setEnabled(selected_row >= 0)

   def search_patient(self):
        """Search patients with filters"""
        try:
            keyword = self.search_input.text().strip().lower()
            age_range = self.age_range.currentText()
            gender = self.gender_filter.currentText()
            
            conn = create_connection()
            c = conn.cursor()
            
            query = "SELECT * FROM patients WHERE 1=1"
            params = []
            
            if keyword:
                query += """ AND (LOWER(name) LIKE ? 
                            OR id_number LIKE ? 
                            OR phone LIKE ?)"""
                params.extend([f'%{keyword}%'] * 3)
                
            if age_range != 'All':
                min_age, max_age = map(int, age_range.replace('Above ', '').split('-'))
                min_date = (datetime.now() - timedelta(days=max_age*365)).strftime('%Y-%m-%d')
                max_date = (datetime.now() - timedelta(days=min_age*365)).strftime('%Y-%m-%d')
                query += " AND dob BETWEEN ? AND ?"
                params.extend([min_date, max_date])
                
            if gender != 'All':
                query += " AND gender = ?"
                params.append(gender)
                
            query += " ORDER BY id DESC"
            
            c.execute(query, params)
            patients = c.fetchall()
            
            self.current_page = 1
            self.total_patients = len(patients)
            self.update_pagination()
            self.display_page(patients)
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Search error: {str(e)}')
        finally:
            conn.close()

   def change_page_size(self):
        self.current_page = 1
        self.search_patient()

   def change_page(self, delta):
        self.current_page += delta
        self.search_patient()

   def update_pagination(self):
        page_size = int(self.page_size.currentText())
        total_pages = (self.total_patients + page_size - 1) // page_size
        
        self.page_label.setText(f"Page {self.current_page}/{total_pages}")
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < total_pages)

   def display_page(self, patients):
        page_size = int(self.page_size.currentText())
        start = (self.current_page - 1) * page_size
        end = start + page_size
        
        page_data = patients[start:end]
        
        self.patient_table.setRowCount(0)
        for row, patient in enumerate(page_data):
            self.patient_table.insertRow(row)
            for col, value in enumerate(patient):
                if col == 2:  # Format date
                    value = self.format_date(value)
                self.patient_table.setItem(row, col, QTableWidgetItem(str(value)))
                
        self.patient_table.clearSelection()

   def export_diagnosis(self):
        """Export current diagnosis to PDF report"""
        if not hasattr(self, 'current_image'):
            QMessageBox.warning(self, 'Error', 'No data to export!')
            return
            
        try:
            patient_data = self.patient_select.currentText().split(' - ')
            filename = QFileDialog.getSaveFileName(
                self,
                "Save PDF Report",
                f"BaoCao_{patient_data[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "PDF Files (*.pdf)"
            )[0]
            
            if filename:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
                
                doc = SimpleDocTemplate(filename, pagesize=A4)
                styles = getSampleStyleSheet()
                story = []
                
                # Title
                story.append(Paragraph("DIAGNOSIS REPORT", styles['Title']))
                story.append(Spacer(1, 20))
                
                # Patient info
                story.append(Paragraph(f"Patient: {patient_data[0]}", styles['Heading2']))
                story.append(Paragraph(f"ID/Passport: {patient_data[2]}", styles['Normal']))
                story.append(Paragraph(f"Phone: {patient_data[1]}", styles['Normal']))
                story.append(Spacer(1, 20))
                
                # Results
                story.append(Paragraph("Detection Results:", styles['Heading3']))
                story.append(Paragraph(self.detect_result.text(), styles['Normal']))
                story.append(Spacer(1, 10))
                
                story.append(Paragraph("Classification Results:", styles['Heading3']))
                story.append(Paragraph(self.classify_result.text(), styles['Normal']))
                story.append(Spacer(1, 10))
                
                # Images
                if hasattr(self, 'current_image'):
                    story.append(Image(self.current_image, width=400, height=300))
                if hasattr(self, 'current_processed_image'):
                    story.append(Image(self.current_processed_image, width=400, height=300))
                    
                # Notes
                if self.diagnosis_notes.toPlainText():
                    story.append(Paragraph("Notes:", styles['Heading3']))
                    story.append(Paragraph(self.diagnosis_notes.toPlainText(), styles['Normal']))
                    
                doc.build(story)
                QMessageBox.information(self, 'Success', 'Diagnosis report exported successfully!')
                
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export diagnosis: {str(e)}')

   def export_excel(self):
        try:
            filename = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                f"DanhSachBenhNhan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )[0]

            if not filename:
                return

            conn = create_connection()
            c = conn.cursor()
            
            # Query data
            c.execute("""
                SELECT 
                    p.id,
                    p.name,
                    p.gender,
                    p.dob,
                    p.phone,
                    p.id_number,
                    p.address,
                    p.email,
                    d.diagnosis_date,
                    d.detection_result,
                    d.classification_result,
                    d.notes
                FROM patients p
                LEFT JOIN diagnoses d ON p.id = d.patient_id
                ORDER BY p.id DESC
            """)
            data = c.fetchall()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Patient List"

            # Headers
            headers = [
                'ID', 'Name', 'Gender', 'Date of Birth', 'Phone', 
                'ID/Passport', 'Address', 'Email', 'Diagnosis Date',
                'Detection Result', 'Classification Result', 'Notes'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

            # Data rows
            for row, record in enumerate(data, 2):
                for col, value in enumerate(record, 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(value, str) and value.startswith('202'):  # Format dates
                        try:
                            date = datetime.strptime(value, '%Y-%m-%d')
                            cell.value = date
                            cell.number_format = 'DD/MM/YYYY'
                        except:
                            cell.value = value
                    else:
                        cell.value = value

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Add statistics sheet
            ws_stats = wb.create_sheet("Statistics")
            
            # Get statistics data
            total_patients = len(set(x[0] for x in data))
            total_diagnoses = len([x for x in data if x[8]])  # Count non-null diagnosis dates
            
            benign_count = len([x for x in data if x[10] and 'Benign' in x[10]])
            malignant_count = len([x for x in data if x[10] and 'Malignan' in x[10]])
            
            stats_data = [
                ['Total Patients:', total_patients],
                ['Total Diagnoses:', total_diagnoses],
                ['Benign Count:', benign_count],
                ['Malignant Count:', malignant_count],
                ['Benign Percentage:', f"{(benign_count/total_diagnoses*100):.1f}%" if total_diagnoses else "0%"],
                ['Malignant Percentage:', f"{(malignant_count/total_diagnoses*100):.1f}%" if total_diagnoses else "0%"]
            ]

            for row, (label, value) in enumerate(stats_data, 1):
                ws_stats.cell(row=row, column=1).value = label
                ws_stats.cell(row=row, column=2).value = value

            # Save workbook
            wb.save(filename)
            QMessageBox.information(self, 'Success', 'Excel file exported successfully!')

        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export Excel: {str(e)}')
        finally:
            conn.close()

   def save_diagnosis(self):
        """Save diagnosis results to database"""
        if not hasattr(self, 'current_image'):
            QMessageBox.warning(self, 'Warning', 'Please select an image!')
            return
            
        patient_id = self.patient_select.currentData()
        if not patient_id:
            QMessageBox.warning(self, 'Warning', 'Please select a patient!')
            return
            
        try:
            # Copy image to diagnoses folder
            filename = f'diagnosis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg'
            image_path = os.path.join(self.diagnoses_dir, filename)
            shutil.copy2(self.current_image, image_path)
            
            # Get processed image path if exists
            processed_path = getattr(self, 'current_processed_image', None)
            
            conn = create_connection()
            c = conn.cursor()
            c.execute("""
                INSERT INTO diagnoses 
                (patient_id, diagnosis_date, image_path, processed_image_path,
                detection_result, classification_result, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                patient_id,
                datetime.now().strftime('%Y-%m-%d'),
                image_path,
                processed_path,
                self.detect_result.text(),
                self.classify_result.text(),
                self.diagnosis_notes.toPlainText()
            ))
            
            conn.commit()
            self.update_statistics()
            QMessageBox.information(self, 'Success', 'Diagnosis results saved successfully!')
            self.clear_image()
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save diagnosis results: {str(e)}')
        finally:
            conn.close()

   def load_data(self):
       try:
           self.search_patient()
           self.load_patients()
           
           if self.role == 'admin':
               self.load_users()
               
           self.update_statistics()
           self.status_bar.showMessage('Ready', 3000)
           
       except Exception as e:
           QMessageBox.critical(self, 'Error', f'Failed to load data: {str(e)}')

   def load_patients(self):
        """Load patients into table and combobox"""
        try:
            conn = create_connection()
            c = conn.cursor()
            c.execute("SELECT * FROM patients ORDER BY id DESC")
            patients = c.fetchall()
            
            # Update table
            self.total_patients = len(patients)
            self.update_pagination()
            self.display_page(patients)
            
            # Update diagnosis combobox if exists
            if hasattr(self, 'patient_select'):
                self.patient_select.clear()
                self.patient_select.addItem("Select Patient...")
                for patient in patients:
                    self.patient_select.addItem(
                        f"{patient[1]} - {patient[6]} - {patient[4]}", 
                        patient[0]
                    )
                    
        except Exception as e:
            print(f"Error loading patients: {e}")
        finally:
            conn.close()

#    def capture_chart(self, chart_view):
#         """Capture chart as image for reports"""
#         try:
#             pixmap = chart_view.grab()
#             temp_path = os.path.join(self.reports_dir, "temp_chart.png")
#             pixmap.save(temp_path)
#             return temp_path
#         except Exception as e:
#             print(f"Error capturing chart: {e}")
#             return None

   def add_patient(self):
        try:
            data = {
                'name': self.patient_name.text().strip(),
                'dob': self.patient_dob.date().toString("yyyy-MM-dd"),
                'gender': self.patient_gender.currentText(),
                'id_number': self.patient_id_number.text().strip(),
                'address': self.patient_address.toPlainText().strip(),
                'phone': self.patient_phone.text().strip(),
                'email': self.patient_email.text().strip()
            }
            
            if not all([data['name'], data['phone'], data['id_number']]):
                QMessageBox.warning(self, 'Error', 'Please enter all required fields!')
                return
                
            conn = create_connection()
            c = conn.cursor()
            
            c.execute("""
                INSERT INTO patients 
                (name, dob, gender, id_number, address, phone, email)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, tuple(data.values()))
            
            conn.commit()
            
            self.load_patients()  
            self.status_bar.showMessage('Patient added successfully', 3000)
            QMessageBox.information(self, 'Success', 'Patient added successfully!')
            
            self.clear_form()
            self.search_patient()
            
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint' in str(e):
                QMessageBox.warning(self, 'Error', 'Phone number or ID already exists!')
            else:
                QMessageBox.critical(self, 'Error', f'Database error: {str(e)}')
        finally:
            conn.close()

   def update_patient(self):
        row = self.patient_table.currentRow()
        if row < 0:
            return
            
        try:
            patient_id = self.patient_table.item(row, 0).text()
            
            data = {
                'name': self.patient_name.text().strip(),
                'dob': self.patient_dob.date().toString("yyyy-MM-dd"), 
                'gender': self.patient_gender.currentText(),
                'id_number': self.patient_id_number.text().strip(),
                'address': self.patient_address.toPlainText().strip(),
                'phone': self.patient_phone.text().strip(),
                'email': self.patient_email.text().strip(),
                'id': patient_id
            }
            
            if not all([data['name'], data['phone'], data['id_number']]):
                QMessageBox.warning(self, 'Error', 'Please enter all required fields!')
                return
                
            conn = create_connection()
            c = conn.cursor()
            
            c.execute("""
                UPDATE patients SET
                name=?, dob=?, gender=?, id_number=?,
                address=?, phone=?, email=?
                WHERE id=?
            """, tuple(data.values()))
            
            conn.commit()
            
            self.load_patients()
            self.status_bar.showMessage('Patient information updated successfully', 3000)
            QMessageBox.information(self, 'Success', 'Patient information updated successfully!')
            
            self.clear_form()
            self.search_patient()
            
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint' in str(e):
                QMessageBox.warning(self, 'Error', 'Phone number or ID already exists!')
            else:
                QMessageBox.critical(self, 'Error', f'Database error: {str(e)}')
        finally:
            conn.close()

   def delete_patient(self):
        row = self.patient_table.currentRow()
        if row < 0:
            return
            
        try:
            patient_id = self.patient_table.item(row, 0).text()
            patient_name = self.patient_table.item(row, 1).text()
            
            reply = QMessageBox.question(
                self, 'Confirmation',
                f'Are you sure you want to delete patient {patient_name}?\n' +
                'All related data will be deleted!',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                conn = create_connection()
                c = conn.cursor()
                
                # Xóa chẩn đoán trước
                c.execute("DELETE FROM diagnoses WHERE patient_id=?", (patient_id,))
                
                # Xóa bệnh nhân 
                c.execute("DELETE FROM patients WHERE id=?", (patient_id,))
                conn.commit()
                
                # Xóa các file ảnh
                self.cleanup_patient_images(patient_id)
                
                self.load_patients()
                self.status_bar.showMessage('Patient deleted successfully', 3000)
                QMessageBox.information(self, 'Success', 'Patient deleted successfully!')
                
                self.clear_form()
                self.search_patient()
                
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Database error: {str(e)}')
        finally:
            conn.close()

   def filter_patients(self):
        """Filter patients in diagnosis tab combobox"""
        search = self.patient_search.text().lower()
        self.patient_select.clear()
        
        try:
            conn = create_connection()
            c = conn.cursor()
            c.execute("""
                SELECT id, name, phone, id_number FROM patients 
                WHERE LOWER(name) LIKE ? OR phone LIKE ? OR id_number LIKE ?
            """, (f'%{search}%', f'%{search}%', f'%{search}%'))
            
            for id_, name, phone, id_num in c.fetchall():
                self.patient_select.addItem(f"{name} - {phone} - {id_num}", id_)
                
        except Exception as e:
            print(f"Error filtering patients: {e}")
        finally:
            conn.close()

   def cleanup_patient_images(self, patient_id):
        try:
            conn = create_connection()
            c = conn.cursor()
            c.execute("SELECT image_path, processed_image_path FROM diagnoses WHERE patient_id=?", (patient_id,))
            images = c.fetchall()
            
            for image_path, processed_path in images:
                if image_path and os.path.exists(image_path):
                    os.remove(image_path)
                if processed_path and os.path.exists(processed_path):
                    os.remove(processed_path)
                    
        except Exception as e:
            print(f"Error cleaning up images: {e}")
        finally:
            conn.close()

   def create_diagnosis_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # Split into left/right panels
        main_layout = QHBoxLayout()

        # Left panel - Image
        left_widget = QWidget()
        left_layout = QVBoxLayout()

        # Image upload area
        upload_group = QGroupBox("Ultrasound Image")
        upload_layout = QVBoxLayout()

        self.image_label = QLabel()
        self.image_label.setMinimumSize(600, 400)
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setStyleSheet("""
            QLabel {
                border: 2px dashed #bdc3c7;
                border-radius: 8px;
                background: #f8f9fa;
            }
        """)
        self.image_label.setText("Drag and drop an image here or click to select")
        self.image_label.setAcceptDrops(True)

        # Image buttons
        btn_layout = QHBoxLayout()
        upload_btn = QPushButton("Select Image")
        upload_btn.setIcon(QIcon("./Code_app/assets/upload.png"))
        upload_btn.clicked.connect(self.upload_image)

        clear_btn = QPushButton("Clear Image")
        clear_btn.setIcon(QIcon("./Code_app/assets/clear.png"))
        clear_btn.clicked.connect(self.clear_image)

        btn_layout.addWidget(upload_btn)
        btn_layout.addWidget(clear_btn)

        upload_layout.addWidget(self.image_label)
        upload_layout.addLayout(btn_layout)
        upload_group.setLayout(upload_layout)
        left_layout.addWidget(upload_group)

        # Image info
        info_group = QGroupBox("Image Info")
        info_layout = QVBoxLayout()
        self.image_info_label = QLabel()
        info_layout.addWidget(self.image_info_label)
        info_group.setLayout(info_layout)
        left_layout.addWidget(info_group)

        left_widget.setLayout(left_layout)
        # Right panel
        right_widget = QWidget()
        right_layout = QVBoxLayout()

        # Patient selection
        patient_group = QGroupBox("Patient Selection")
        patient_layout = QVBoxLayout()
        
        search_layout = QHBoxLayout()
        self.patient_search = QLineEdit()
        self.patient_search.setPlaceholderText("Search patient...")
        self.patient_search.textChanged.connect(self.filter_patients)
        
        self.patient_select = QComboBox()
        self.patient_select.setMinimumWidth(300)
        self.load_patients()
        
        search_layout.addWidget(self.patient_search)
        search_layout.addWidget(self.patient_select)
        
        self.patient_info = QLabel()
        
        patient_layout.addLayout(search_layout)
        patient_layout.addWidget(self.patient_info)
        patient_group.setLayout(patient_layout)
        right_layout.addWidget(patient_group)

        # AI Results
        results_group = QGroupBox("AI Results")
        results_layout = QVBoxLayout()

        # Detection
        detect_widget = QWidget()
        detect_layout = QVBoxLayout()
        
        detect_label = QLabel("DETECTION:")
        detect_label.setStyleSheet("font-weight: bold;")
        
        self.detect_progress = QProgressBar()
        self.detect_progress.setTextVisible(False)
        
        self.detect_result = QLabel("No result")
        self.detect_result.setStyleSheet("""
            background: #f8f9fa;
            padding: 10px;
            border: 1px solid #dee2e6;
            border-radius: 4px;
        """)
        
        detect_btn = QPushButton("Run Detect")
        detect_btn.clicked.connect(self.run_detection)
        detect_btn.setIcon(QIcon("./Code_app/assets/detect.png"))
        
        detect_layout.addWidget(detect_label)
        detect_layout.addWidget(self.detect_progress)
        detect_layout.addWidget(self.detect_result)
        detect_layout.addWidget(detect_btn)
        detect_widget.setLayout(detect_layout)

        # Classification  
        classify_widget = QWidget()
        classify_layout = QVBoxLayout()
        
        classify_label = QLabel("CLASSIFICATION:")
        classify_label.setStyleSheet("font-weight: bold;")
        
        self.classify_progress = QProgressBar()
        self.classify_progress.setTextVisible(False)
        
        self.classify_result = QLabel("No result")  
        self.classify_result.setStyleSheet("""
            background: #f8f9fa;
            padding: 10px;
            border: 1px solid #dee2e6;
            border-radius: 4px;
        """)
        
        classify_btn = QPushButton("Run Classify")
        classify_btn.clicked.connect(self.run_classification)
        classify_btn.setIcon(QIcon("./Code_app/assets/classify.png"))
        
        classify_layout.addWidget(classify_label)
        classify_layout.addWidget(self.classify_progress)
        classify_layout.addWidget(self.classify_result)
        classify_layout.addWidget(classify_btn)
        classify_widget.setLayout(classify_layout)

        results_layout.addWidget(detect_widget)
        results_layout.addWidget(classify_widget)
        results_group.setLayout(results_layout)
        right_layout.addWidget(results_group)
        # Notes section
        notes_group = QGroupBox("Notes")
        notes_layout = QVBoxLayout()
        self.diagnosis_notes = QTextEdit()
        self.diagnosis_notes.setPlaceholderText("Enter diagnosis notes...")
        notes_layout.addWidget(self.diagnosis_notes)
        notes_group.setLayout(notes_layout)
        right_layout.addWidget(notes_group)

        # Save & Export buttons
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("Save Results")
        save_btn.setIcon(QIcon("./Code_app/assets/save.png"))
        save_btn.clicked.connect(self.save_diagnosis)
        
        export_btn = QPushButton("Export Report")
        export_btn.setIcon(QIcon("./Code_app/assets/export.png"))
        export_btn.clicked.connect(self.export_diagnosis)
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(export_btn)
        right_layout.addLayout(btn_layout)

        right_widget.setLayout(right_layout)

        # Add panels to main layout
        main_layout.addWidget(left_widget, stretch=6)
        main_layout.addWidget(right_widget, stretch=4)
        layout.addLayout(main_layout)

        tab.setLayout(layout)
        return tab

   def run_detection(self):
        if not hasattr(self, 'current_image'):
            QMessageBox.warning(self, 'Error', 'Please select an image!')
            return
            
        try:
            self.detect_progress.setValue(0)
            self.detect_progress.setMaximum(0)
            
            # Run YOLO detection
            results = self.yolo_model.predict(self.current_image)
            boxes = results[0].boxes
            
            self.detect_progress.setMaximum(100)
            self.detect_progress.setValue(50)
            
            if len(boxes) > 0:
                # Process image with bounding boxes
                processed_img = self.process_detection_image(boxes)
                if processed_img:
                    pixmap = QPixmap(processed_img)
                    self.image_label.setPixmap(pixmap.scaled(
                        self.image_label.size(),
                        Qt.KeepAspectRatio,
                        Qt.SmoothTransformation
                    ))
                    
                    self.current_processed_image = processed_img
                    
                self.detect_result.setText(f"Detected {len(boxes)} objects")
                self.detect_result.setStyleSheet("color: #e74c3c;")
            else:
                self.detect_result.setText("No objects detected")
                self.detect_result.setStyleSheet("color: #27ae60;")
                
            self.detect_progress.setValue(100)
            self.status_bar.showMessage('Detection completed', 3000)
                
        except Exception as e:
            self.detect_progress.setValue(0)
            QMessageBox.critical(self, 'Error', f'Detect error: {str(e)}')

   def process_detection_image(self, boxes):
        try:
            img = cv2.imread(self.current_image)
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            
            # Draw boxes and confidence scores
            for box in boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                conf = float(box.conf[0])
                
                color = (0, int(255 * conf), 0)
                cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
                
                text = f"{conf:.2f}"
                font = cv2.FONT_HERSHEY_SIMPLEX
                cv2.putText(img, text, (x1, y1-10), font, 0.6, color, 2)
                
            filename = f'processed_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg'
            save_path = os.path.join(self.processed_dir, filename)
            cv2.imwrite(save_path, cv2.cvtColor(img, cv2.COLOR_RGB2BGR))
            
            return save_path
                
        except Exception as e:
            raise Exception(f'Process error: {str(e)}')

   def run_classification(self):
        if not hasattr(self, 'current_image'):
            QMessageBox.warning(self, 'Error', 'Please select an image!')
            return
                
        try:
            self.classify_progress.setValue(0)
            self.classify_progress.setMaximum(0)
            
            # Preprocess image
            image = tf.keras.utils.load_img(
                self.current_image,
                target_size=(224, 224)
            )
            img_array = tf.keras.utils.img_to_array(image)
            img_array = img_array / 255.0
            img_array = tf.expand_dims(img_array, 0)
            
            self.classify_progress.setMaximum(100)
            self.classify_progress.setValue(50)
            
            # Make prediction
            predictions = self.keras_model.predict(img_array)
            class_names = ['Benign', 'Malignant', 'Normal']
            predicted_class = np.argmax(predictions[0])
            confidence = predictions[0][predicted_class]
            
            # Map results
            result_map = {
                'Benign': 'Benign',
                'Malignant': 'Malignan',
                'Normal': 'Normal'
            }
            result = result_map[class_names[predicted_class]]
            
            # Display results with colors
            colors = {
                'Benign': '#27ae60',
                'Malignan': '#e74c3c', 
                'Normal': '#2980b9'
            }
            self.classify_result.setText(f"Result: {result} ({confidence:.2%})")
            self.classify_result.setStyleSheet(f"color: {colors[result]};")
            
            self.classify_progress.setValue(100)
            self.status_bar.showMessage('Classification completed', 3000)
                
        except Exception as e:
            self.classify_progress.setValue(0)
            QMessageBox.critical(self, 'Error', f'Classify error: {str(e)}')

   def patient_table_clicked(self):
        """Handle patient table row selection"""
        row = self.patient_table.currentRow()
        if row >= 0:
            # Get patient data from selected row
            self.patient_name.setText(self.patient_table.item(row, 1).text())
            self.patient_dob.setDate(QDate.fromString(self.patient_table.item(row, 2).text(), "dd/MM/yyyy"))
            self.patient_gender.setCurrentText(self.patient_table.item(row, 3).text())
            self.patient_id_number.setText(self.patient_table.item(row, 4).text())
            self.patient_address.setText(self.patient_table.item(row, 5).text())
            self.patient_phone.setText(self.patient_table.item(row, 6).text())
            self.patient_email.setText(self.patient_table.item(row, 7).text())
            
            self.validate_form()

   def create_reports_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Statistics Dashboard
        stats_group = QGroupBox("Statistics Dashboard")
        stats_layout = QGridLayout()
        
        # KPI Indicators
        self.total_patients_label = QLabel("0")
        self.total_diagnoses_label = QLabel("0")
        self.benign_rate_label = QLabel("0%")
        self.malignant_rate_label = QLabel("0%")
        
        # Add KPI cards
        stats_layout.addWidget(self.create_kpi_card("Total Patients", self.total_patients_label), 0, 0)
        stats_layout.addWidget(self.create_kpi_card("Total Diagnoses", self.total_diagnoses_label), 0, 1)
        stats_layout.addWidget(self.create_kpi_card("Benign Rate", self.benign_rate_label), 0, 2)
        stats_layout.addWidget(self.create_kpi_card("Malignant Rate", self.malignant_rate_label), 0, 3)
        
        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)
        
        # Charts Section
        charts_layout = QHBoxLayout()
        
        # Monthly trend chart
        trend_group = QGroupBox("Monthly Diagnosis Trend")
        trend_layout = QVBoxLayout()
        self.trend_chart = QChartView()
        trend_layout.addWidget(self.trend_chart)
        trend_group.setLayout(trend_layout)
        charts_layout.addWidget(trend_group)
        
        # Distribution chart
        dist_group = QGroupBox("Distribution Chart")
        dist_layout = QVBoxLayout()
        self.dist_chart = QChartView()
        dist_layout.addWidget(self.dist_chart)
        dist_group.setLayout(dist_layout)
        charts_layout.addWidget(dist_group)
        
        layout.addLayout(charts_layout)
        
        # Report Generation 
        report_group = QGroupBox("Report Generation")
        report_layout = QGridLayout()
        
        # Date range
        self.report_date_from = QDateEdit()
        self.report_date_from.setCalendarPopup(True)
        self.report_date_from.setDate(QDate.currentDate().addMonths(-1))
        
        self.report_date_to = QDateEdit()
        self.report_date_to.setCalendarPopup(True)
        self.report_date_to.setDate(QDate.currentDate())
        
        report_layout.addWidget(QLabel("From:"), 0, 0)
        report_layout.addWidget(self.report_date_from, 0, 1)
        report_layout.addWidget(QLabel("To:"), 0, 2)
        report_layout.addWidget(self.report_date_to, 0, 3)
        
        # Report options
        self.include_charts = QCheckBox("Include Charts")
        self.include_charts.setChecked(True)
        self.include_patient_info = QCheckBox("Include Patient Info")
        self.include_patient_info.setChecked(True)
        self.include_images = QCheckBox("Include Images")
        self.include_images.setChecked(True)
        
        options_layout = QVBoxLayout()
        options_layout.addWidget(self.include_charts)
        options_layout.addWidget(self.include_patient_info)
        options_layout.addWidget(self.include_images)
        
        report_layout.addLayout(options_layout, 1, 0, 2, 4)
        
        # Generate button
        generate_btn = QPushButton("Generate Report")
        generate_btn.setIcon(QIcon("./Code_app/assets/report.png"))
        generate_btn.clicked.connect(self.generate_report) 
        report_layout.addWidget(generate_btn, 3, 0, 1, 4, Qt.AlignCenter)
        
        report_group.setLayout(report_layout)
        layout.addWidget(report_group)
        
        tab.setLayout(layout)
        return tab

   def create_kpi_card(self, title, value_label):
        card = QWidget()
        layout = QVBoxLayout()
        
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            font-size: 14px;
            color: #7f8c8d;
        """)
        
        value_label.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #2980b9;
        """)
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.setAlignment(Qt.AlignCenter)
        
        card.setLayout(layout)
        card.setStyleSheet("""
            QWidget {
                background: white;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        
        return card

   def update_statistics(self):
        try:
            conn = create_connection()
            c = conn.cursor()
            
            # Cập nhật tổng số bệnh nhân
            c.execute("SELECT COUNT(DISTINCT id) FROM patients")
            self.total_patients_label.setText(str(c.fetchone()[0]))
            
            # Cập nhật số ca chẩn đoán
            c.execute("SELECT COUNT(*) FROM diagnoses") 
            self.total_diagnoses_label.setText(str(c.fetchone()[0]))
            
            # Cập nhật tỷ lệ kết quả
            c.execute("""
                SELECT classification_result, COUNT(*) as count
                FROM diagnoses 
                GROUP BY classification_result
            """)
            results = dict(c.fetchall())
            
            total = sum(results.values()) if results else 0
            if total > 0:
                benign = sum(v for k,v in results.items() if 'Benign' in k)
                malignant = sum(v for k,v in results.items() if 'Malignan' in k)
                
                self.benign_rate_label.setText(f"{benign/total*100:.1f}%")
                self.malignant_rate_label.setText(f"{malignant/total*100:.1f}%")
                
            # Cập nhật biểu đồ
            self.update_trend_chart()
            self.update_distribution_chart()
            
        except Exception as e:
            print(f"Error updating statistics: {e}")
        finally:
            conn.close()

   def update_trend_chart(self):
        try:
            conn = create_connection()
            c = conn.cursor()
            
            c.execute("""
                SELECT 
                    strftime('%Y-%m', diagnosis_date) as month,
                    COUNT(*) as count,
                    SUM(CASE WHEN classification_result LIKE '%Benign%' THEN 1 ELSE 0 END) as benign,
                    SUM(CASE WHEN classification_result LIKE '%Malignan%' THEN 1 ELSE 0 END) as malignant
                FROM diagnoses 
                WHERE diagnosis_date >= date('now', '-12 months')
                GROUP BY month
                ORDER BY month
            """)
            data = c.fetchall()
            
            series = QBarSeries()
            benign_set = QBarSet("Benign")
            malignant_set = QBarSet("Malignan")
            
            categories = []
            for month, total, benign, malignant in data:
                categories.append(month)
                benign_set.append(benign)
                malignant_set.append(malignant)
            
            series.append(benign_set)
            series.append(malignant_set)
            
            chart = QChart()
            chart.addSeries(series)
            chart.setTitle("Number of Diagnoses by Month")
            chart.setAnimationOptions(QChart.SeriesAnimations)
            
            axis_x = QBarCategoryAxis()
            axis_x.append(categories)
            chart.addAxis(axis_x, Qt.AlignBottom)
            series.attachAxis(axis_x)
            
            axis_y = QValueAxis()
            chart.addAxis(axis_y, Qt.AlignLeft)
            series.attachAxis(axis_y)
            
            self.trend_chart.setChart(chart)
            
        except Exception as e:
            print(f"Error updating trend chart: {e}")
        finally:
            conn.close()
   def update_distribution_chart(self):
        try:
            conn = create_connection()
            c = conn.cursor()
            
            c.execute("""
                SELECT 
                    CASE 
                        WHEN classification_result LIKE '%Benign%' THEN 'Benign'
                        WHEN classification_result LIKE '%Malignan%' THEN 'Malignan' 
                        ELSE 'Normal'
                    END as type,
                    COUNT(*) as count
                FROM diagnoses
                GROUP BY type
            """)
            data = c.fetchall()
            
            series = QPieSeries()
            colors = ['#27ae60', '#e74c3c', '#3498db']
            
            for i, (type_, count) in enumerate(data):
                slice_ = series.append(f"{type_} ({count})", count)
                slice_.setLabelVisible(True)
                slice_.setBrush(QColor(colors[i]))
                
            chart = QChart()
            chart.addSeries(series)
            chart.setTitle("Distribution of Diagnosis Results")
            chart.setAnimationOptions(QChart.SeriesAnimations)
            chart.legend().setAlignment(Qt.AlignRight)
            
            self.dist_chart.setChart(chart)
            
        except Exception as e:
            print(f"Error updating distribution chart: {e}")
        finally:
            conn.close()

   def create_admin_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # Users Management
        users_group = QGroupBox("User Management")
        users_layout = QVBoxLayout()
        
        # Users table
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)
        self.users_table.setHorizontalHeaderLabels([
            'ID', 'Username', 'Role', 'Created At'
        ])
        
        header = self.users_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # ID column
        self.users_table.setColumnWidth(0, 50)
        
        self.users_table.verticalHeader().setVisible(False)
        self.users_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.users_table.setSelectionMode(QTableWidget.SingleSelection)
        self.users_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # User management buttons
        btn_layout = QHBoxLayout()
        
        add_user_btn = QPushButton("Add User")
        add_user_btn.setIcon(QIcon("./Code_app/assets/add_user.png"))
        add_user_btn.clicked.connect(self.add_user)
        
        delete_user_btn = QPushButton("Delete User")
        delete_user_btn.setIcon(QIcon("./Code_app/assets/delete_user.png"))
        delete_user_btn.clicked.connect(self.delete_user)
        
        btn_layout.addWidget(add_user_btn)
        btn_layout.addWidget(delete_user_btn)
        
        users_layout.addWidget(self.users_table)
        users_layout.addLayout(btn_layout)
        users_group.setLayout(users_layout)
        layout.addWidget(users_group)

        # System settings
        settings_group = QGroupBox("System Settings")
        settings_layout = QVBoxLayout()
        
        # Backup schedule
        backup_layout = QHBoxLayout()
        backup_layout.addWidget(QLabel("Backup Schedule:"))
        
        self.backup_schedule = QComboBox()
        self.backup_schedule.addItems(['Never', 'Daily', 'Weekly', 'Monthly'])
        backup_layout.addWidget(self.backup_schedule)
        
        settings_layout.addLayout(backup_layout)
        
        # Model settings
        model_layout = QGridLayout()
        model_layout.addWidget(QLabel("Detect Threshold:"), 0, 0)
        
        self.detect_threshold = QLineEdit()
        self.detect_threshold.setPlaceholderText("0.5")
        model_layout.addWidget(self.detect_threshold, 0, 1)
        
        model_layout.addWidget(QLabel("Classify Threshold:"), 1, 0)    
        
        self.classify_threshold = QLineEdit()
        self.classify_threshold.setPlaceholderText("0.7")
        model_layout.addWidget(self.classify_threshold, 1, 1)
        
        settings_layout.addLayout(model_layout)
        
        # Save settings button
        save_settings_btn = QPushButton("Save Settings")
        save_settings_btn.clicked.connect(self.save_settings)
        settings_layout.addWidget(save_settings_btn)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        tab.setLayout(layout)
        return tab

   def add_user(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add User")
        layout = QFormLayout()
        
        username = QLineEdit()
        password = QLineEdit()
        password.setEchoMode(QLineEdit.Password)
        
        role = QComboBox()
        role.addItems(['doctor', 'admin'])
        
        layout.addRow("Username:", username)
        layout.addRow("Password:", password) 
        layout.addRow("Role:", role)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        
        layout.addRow(buttons)
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            try:
                conn = create_connection()
                c = conn.cursor()
                
                c.execute("""
                    INSERT INTO users (username, password, role)
                    VALUES (?, ?, ?)
                """, (username.text(), password.text(), role.currentText()))
                
                conn.commit()
                self.load_users()
                
                QMessageBox.information(self, 'Success', 'User added successfully!')
                
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, 'Error', 'Username already exists!')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Failed to add user: {str(e)}')
            finally:
                conn.close()
   def delete_user(self):
        row = self.users_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Error', 'Please select a user to delete!')
            return

        try:
            user_id = self.users_table.item(row, 0).text()
            username = self.users_table.item(row, 1).text()
            
            if username == 'admin':
                QMessageBox.warning(self, 'Error', 'You cannot delete the admin account!')
                return
                
            reply = QMessageBox.question(
                self, 'Confirmation',
                f'Are you sure you want to delete user {username}?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                conn = create_connection()
                c = conn.cursor()
                c.execute("DELETE FROM users WHERE id=?", (user_id,))
                conn.commit()
                
                self.load_users()
                QMessageBox.information(self, 'Success', 'User deleted successfully!')
                
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to delete user: {str(e)}')
        finally:
            conn.close()

   def save_settings(self):
        try:
            # Save detect threshold
            detect_threshold = float(self.detect_threshold.text() or 0.5)
            if not 0 <= detect_threshold <= 1:
                raise ValueError("Detect threshold must be between 0-1")
                
            # Save classify threshold
            classify_threshold = float(self.classify_threshold.text() or 0.7) 
            if not 0 <= classify_threshold <= 1:
                raise ValueError("Classify threshold must be between 0-1")
                
            # Save backup schedule
            backup_schedule = self.backup_schedule.currentText()
            
            # Save settings to file
            settings = {
                'detect_threshold': detect_threshold,
                'classify_threshold': classify_threshold,
                'backup_schedule': backup_schedule
            }
            
            with open('./Code_app/settings.json', 'w') as f:
                json.dump(settings, f)
                
            QMessageBox.information(self, 'Success', 'Settings saved successfully!')
            
        except ValueError as e:
            QMessageBox.warning(self, 'Error', str(e))
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save settings: {str(e)}')

   def load_settings(self):
        try:
            if os.path.exists('./Code_app/settings.json'):
                with open('./Code_app/settings.json') as f:
                    settings = json.load(f)
                    
                self.detect_threshold.setText(str(settings.get('detect_threshold', '')))
                self.classify_threshold.setText(str(settings.get('classify_threshold', '')))
                
                schedule = settings.get('backup_schedule', 'None')
                index = self.backup_schedule.findText(schedule)
                if index >= 0:
                    self.backup_schedule.setCurrentIndex(index)
                    
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load settings: {str(e)}')

   def load_users(self):
        try:
            conn = create_connection()
            c = conn.cursor()
            c.execute("SELECT * FROM users ORDER BY id")
            users = c.fetchall()
                
            self.users_table.setRowCount(0)
            for row, user in enumerate(users):
                self.users_table.insertRow(row)
                for col, value in enumerate(user):
                    self.users_table.setItem(row, col, QTableWidgetItem(str(value)))
                        
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load users: {str(e)}')
        finally:
            conn.close()

   def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

   def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            self.load_image(files[0])

   def load_image(self, filepath):
        try:
            if not os.path.exists(filepath):
                raise FileNotFoundError("File does not exist")
                
            pixmap = QPixmap(filepath)
            if pixmap.isNull():
                raise ValueError("Failed to load image")
                
            # Scale image while maintaining aspect ratio
            scaled_pixmap = pixmap.scaled(
                self.image_label.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            
            self.image_label.setPixmap(scaled_pixmap)
            self.current_image = filepath
            
            # Reset results
            self.detect_result.setText("No result")
            self.classify_result.setText("No result")
            
            # Show image info
            self.show_image_info(filepath)
            self.status_bar.showMessage('Image loaded successfully', 3000)
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load image: {str(e)}')

   def upload_image(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Select Image",
            "",
            "Image Files (*.png *.jpg *.jpeg)"
        )
        if file_name:
            self.load_image(file_name)

   def clear_image(self):
        self.image_label.clear()
        self.image_label.setText("No image")
        
        if hasattr(self, 'current_image'):
            delattr(self, 'current_image')
        if hasattr(self, 'current_processed_image'):
            delattr(self, 'current_processed_image')
            
        self.detect_result.setText("No result")
        self.classify_result.setText("No result")
        self.image_info_label.clear()
        self.diagnosis_notes.clear()
        
        self.status_bar.showMessage('Image cleared', 3000)

   def show_image_info(self, filepath):
        try:
            # Get image dimensions
            img = cv2.imread(filepath)
            height, width = img.shape[:2]
            
            # Get file size
            size = os.path.getsize(filepath) / 1024  # KB
            
            # Get image format
            _, ext = os.path.splitext(filepath)
            
            # Get last modified time
            modified = datetime.fromtimestamp(os.path.getmtime(filepath))
            
            info = f"""
            Image size: {width}x{height}px
            Image format: {ext.upper()[1:]}
            Image size: {size:.1f}KB
            Last modified: {modified.strftime('%d/%m/%Y %H:%M')}
            File: {filepath}
            """
            self.image_info_label.setText(info)
            
        except Exception as e:
            self.image_info_label.setText("Failed to read image information")
            print(f"Error reading image info: {e}")

   def show_model_info(self):
        info = QDialog(self)
        info.setWindowTitle("Model AI Information")
        layout = QVBoxLayout()
        
        metrics = self.model_metrics
        
        # YOLO metrics
        yolo_group = QGroupBox("YOLO Detection Model")
        yolo_layout = QFormLayout()
        
        yolo_metrics = [
            ("Precision:", f"{metrics['detect']['precision']:.2%}"),
            ("Recall:", f"{metrics['detect']['recall']:.2%}"),
            ("mAP@50:", f"{metrics['detect']['map50']:.2%}")
        ]
        
        for label, value in yolo_metrics:
            value_label = QLabel(value)
            value_label.setStyleSheet("font-weight: bold;")
            yolo_layout.addRow(label, value_label)
            
        yolo_group.setLayout(yolo_layout)
        # DenseNet metrics
        dense_group = QGroupBox("DenseNet Classification Model")
        dense_layout = QFormLayout()
        
        dense_metrics = [
            ("Accuracy:", f"{metrics['classify']['accuracy']:.2%}"),
            ("F1-Score:", f"{metrics['classify']['f1']:.2%}"),
            ("AUC:", f"{metrics['classify']['auc']:.2%}")
        ]
        
        for label, value in dense_metrics:
            value_label = QLabel(value)
            value_label.setStyleSheet("font-weight: bold;")
            dense_layout.addRow(label, value_label)
            
        dense_group.setLayout(dense_layout)
        
        # Training params 
        params_group = QGroupBox("Training Parameters")
        params_layout = QFormLayout()
        
        params = [
            ("Epochs:", "100"),
            ("Batch size:", "32"),
            ("Optimizer:", "Adam"),
            ("Learning rate:", "0.001"),
            ("Input size:", "224x224"),
            ("Dataset:", "5000 images")
        ]
        
        for label, value in params:
            value_label = QLabel(value)
            value_label.setStyleSheet("font-family: monospace;")
            params_layout.addRow(label, value_label)
            
        params_group.setLayout(params_layout)
        
        # Add all groups
        layout.addWidget(yolo_group)
        layout.addWidget(dense_group)
        layout.addWidget(params_group)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(info.close)
        layout.addWidget(close_btn)
        
        info.setLayout(layout)
        info.exec_()

   def capture_chart(self, chart_view):
        """Capture chart as temporary image file"""
        try:
            # Tạo thư mục tạm nếu chưa tồn tại
            temp_dir = os.path.join(os.getcwd(), 'temp')
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            # Tạo tên file tạm unique
            temp_filename = f'chart_{int(time.time() * 1000)}.png'
            temp_path = os.path.join(temp_dir, temp_filename)

            # Capture chart trực tiếp và lưu
            pixmap = chart_view.grab()
            pixmap = pixmap.scaled(
                1200, 800,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            
            # Lưu trực tiếp không qua buffer
            if pixmap.save(temp_path, "PNG", quality=100):
                return temp_path
            else:
                print("Cannot save chart image")
                return None

        except Exception as e:
            print(f"Error in capture_chart: {str(e)}")
            return None

   def generate_report(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle 
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.enums import TA_LEFT, TA_CENTER

            # Register font for Vietnamese
            pdfmetrics.registerFont(TTFont('DejaVuSans', './Code_app/assets/fonts/DejaVuSans.ttf'))

            filename = QFileDialog.getSaveFileName(
                self,
                "Save PDF Report",
                f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                "PDF Files (*.pdf)"
            )[0]

            if not filename:
                return

            # Get date range for query
            date_from = self.report_date_from.date().toString("yyyy-MM-dd")
            date_to = self.report_date_to.date().toString("yyyy-MM-dd") 

            # Query data
            conn = create_connection()
            c = conn.cursor()

            c.execute("""
                SELECT 
                    p.name, p.gender, p.dob, p.phone, p.id_number,
                    d.diagnosis_date, d.detection_result,
                    d.classification_result, d.notes,
                    d.image_path, d.processed_image_path
                FROM diagnoses d
                JOIN patients p ON d.patient_id = p.id 
                WHERE d.diagnosis_date BETWEEN ? AND ?
                ORDER BY d.diagnosis_date DESC
            """, (date_from, date_to))

            data = c.fetchall()

            # Create PDF
            doc = SimpleDocTemplate(
                filename,
                pagesize=A4,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30,
                encoding='utf-8'
            )

            # Define styles with Vietnamese support
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='CustomTitle',
                fontName='DejaVuSans',
                fontSize=20,
                alignment=TA_CENTER,
                spaceAfter=30,
                leading=24
            ))
            
            styles.add(ParagraphStyle(
                name='CustomNormal',
                fontName='DejaVuSans',
                fontSize=11,
                leading=14,
                wordWrap='CJK'
            ))

            styles.add(ParagraphStyle(
                name='CustomHeading',
                fontName='DejaVuSans',
                fontSize=14,
                leading=16,
                spaceAfter=12
            ))

            story = []

            # Title
            story.append(Paragraph("BREAST CANCER DIAGNOSIS REPORT", styles['CustomTitle']))
            story.append(Spacer(1, 20))

            # Statistics
            benign_count = len([x for x in data if 'Benign' in x[7]])
            malignant_count = len([x for x in data if 'Malignan' in x[7]])
            total = len(data)

            benign_rate = (benign_count/total*100) if total > 0 else 0
            malignant_rate = (malignant_count/total*100) if total > 0 else 0

            stats_data = [
                ['Total cases:', str(total)],
                ['Time period:', f"{self.format_date(date_from)} - {self.format_date(date_to)}"],
                ['Benign rate:', f"{benign_rate:.1f}%"],
                ['Malignant rate:', f"{malignant_rate:.1f}%"]
            ]

            # Create wrapper function for table cells
            def create_cell(text):
                return Paragraph(text, styles['CustomNormal'])

            stats_table = Table([[create_cell(cell) for cell in row] for row in stats_data], 
                            colWidths=[150, 350])
            
            stats_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story.append(stats_table)
            story.append(Spacer(1, 20))

            # Charts
            if self.include_charts.isChecked() and total > 0:
                story.append(Paragraph("CHARTS", styles['CustomHeading']))
                
                # Trend chart
                trend_img = self.capture_chart(self.trend_chart)
                if trend_img and os.path.exists(trend_img):
                    try:
                        story.append(Image(trend_img, width=500, height=300))
                        story.append(Spacer(1, 10))
                    except Exception as e:
                        print(f"Error adding trend chart: {e}")
                    finally:
                        # Đảm bảo xóa file tạm
                        try:
                            os.remove(trend_img)
                        except:
                            pass

                # Distribution chart  
                dist_img = self.capture_chart(self.dist_chart)
                if dist_img and os.path.exists(dist_img):
                    try:
                        story.append(Image(dist_img, width=500, height=300))
                        story.append(Spacer(1, 20))
                    except Exception as e:
                        print(f"Error adding distribution chart: {e}")
                    finally:
                        # Đảm bảo xóa file tạm
                        try:
                            os.remove(dist_img)
                        except:
                            pass

            # Patient details
            if total > 0:
                for case in data:
                    story.append(PageBreak())
                    story.append(Paragraph("PATIENT DETAILS", styles['CustomHeading']))
                    
                    if self.include_patient_info.isChecked():
                        story.append(Paragraph(f"PATIENT: {case[0]}", styles['CustomNormal']))
                        story.append(Spacer(1, 10))

                        info_data = [
                            ['GENDER:', case[1]],
                            ['DATE OF BIRTH:', self.format_date(case[2])],
                            ['PHONE NUMBER:', case[3]],
                            ['ID NUMBER:', case[4]],
                            ['DIAGNOSIS DATE:', self.format_date(case[5])]
                        ]

                        info_table = Table([[create_cell(cell) for cell in row] for row in info_data],
                                        colWidths=[150, 350])
                        
                        info_table.setStyle(TableStyle([
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('PADDING', (0, 0), (-1, -1), 8),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))

                        story.append(info_table)
                        story.append(Spacer(1, 10))

                    # Results with wrapped text
                    result_data = [
                        ['DETECT RESULT:', case[6]],
                        ['CLASSIFY RESULT:', case[7]],
                        ['NOTE:', case[8] or 'No note']
                    ]

                    result_table = Table([[create_cell(cell) for cell in row] for row in result_data],
                                    colWidths=[150, 350])
                    
                    result_table.setStyle(TableStyle([
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('PADDING', (0, 0), (-1, -1), 8),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ]))

                    story.append(result_table)

                    # Images
                    if self.include_images.isChecked():
                        # Ảnh gốc
                        if case[9] and os.path.exists(case[9]):
                            story.append(Spacer(1, 10))
                            story.append(Paragraph("ORIGINAL IMAGE:", styles['CustomNormal']))
                            story.append(Spacer(1, 5))
                            story.append(Image(case[9], width=400, height=300))
                        
                        # Thêm PageBreak trước ảnh đã xử lý
                        if case[10] and os.path.exists(case[10]):
                            story.append(PageBreak())
                            story.append(Paragraph("PROCESSED IMAGE:", styles['CustomNormal']))
                            story.append(Spacer(1, 5))
                            story.append(Image(case[10], width=400, height=300))

            else:
                story.append(Paragraph("No data in this time period", styles['CustomNormal']))

            # Build PDF
            doc.build(story)
            
            QMessageBox.information(self, 'Success', 'Report generated successfully!')

        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to generate report: {str(e)}')
        finally:
            conn.close()

   def format_date(self, date_str):
        """Format date string from yyyy-MM-dd to dd/MM/yyyy"""
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d')
            return date.strftime('%d/%m/%Y')
        except:
            return date_str

   def closeEvent(self, event):
        """Handle application closing"""
        reply = QMessageBox.question(
            self, 'Confirmation',
            'Are you sure you want to quit?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

if __name__ == '__main__':
   app = QApplication(sys.argv)
   app.setStyle('Fusion')
   
   from database import create_database
   create_database()
   
   window = MainWindow('admin')
   window.show()
   
   sys.exit(app.exec_())